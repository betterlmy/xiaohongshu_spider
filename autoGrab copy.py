import time
from tkinter import N
from appium import webdriver
from appium.webdriver.common.appiumby import AppiumBy
import os
import sys
import json
import openpyxl
from utils import print_log_nowtime

"""
    1. 小红书的控件id竟然会变化!!!!
    2. 真恶心
    3. !w! 表示warning. 出现提示
    4. !e! 表示error.   出现程序错误
    5. 仍需要完成的任务: 邮箱过滤,前端框架,点击更多,
"""
now_time = ""

wait_time = .5

known_xhs_id = []

def print_log(x, end="\n"):
    """重写print_log"""
    print_log_nowtime(x,now_time=now_time,end=end)

def pass_ad(driver):
    """跳过广告"""

    try:
        driver.find_element(by=AppiumBy.ID, value=resource_id["pass"]).click()
        time.sleep(wait_time)
        print_log("跳过广告成功", end="-->")
    except:
        print_log("不存在广告", end="-->")


def get_xml(driver, xml="xml/info.xml"):
    """将页面资源写入到xml文件中"""

    if not os.path.exists("xml/"):
        os.mkdir("xml")
        print_log("xml文件夹不存在,已自动创建")

    if driver is None:
        print_log(f"!e!Driver不存在")
    else:
        with open(xml, "w", encoding="utf-8") as f:
            f.writelines(driver.page_source)
            print_log(f"xml文件成功保存到{xml}", end="-->")


def contain_video(driver):
    """判断是否存在视频,存在返回True,不存在返回False"""

    try:
        driver.find_element(by=AppiumBy.ID, value=resource_id["video"])
        time.sleep(wait_time)
        return True
    except:
        return False


def fresh_page(driver):
    """刷新页面"""

    try:
        # es = driver.find_elements(by=AppiumBy.CLASS_NAME, value='android.widget.TextView')
        # for e in es:
        #     if e.text == "发现":
        #         e.click()

        driver.find_element(by=AppiumBy.ID, value=resource_id['fresh']).click()
        # time.sleep(wait_time)
        print_log("页面刷新成功", end="-->")
        return True
    except:
        print_log("不存在刷新按钮,刷新失败")
        return False


def auto(driver, max_fresh_num=8):
    """自动化主要流程"""

    fresh_num = 0  # 刷新次数
    while contain_video(driver):
        #如果包含了视频,直接fresh
        print_log("发现视频", end="-->")
        if fresh_page(driver):
            fresh_num += 1
            if fresh_num > max_fresh_num:
                print_log("刷新次数过多仍未获取到信息,退出程序")
                return False
        else:
            return False

    print_log("未发现视频,开始获取信息")
    for index in range(4):
        # 每个阶段获取四个博主信息
        try:
            wrong_step = "获取文章控件"  # 出现异常的阶段
            articles = driver.find_elements(
                by=AppiumBy.ID, value=resource_id["article"])
            articles[index].click()

            wrong_step = "点击文章控件"
            print_log(f"进入第{len(info_list)+1}篇文章", end="-->")
            time.sleep(wait_time)
            e2 = driver.find_element(
                by=AppiumBy.ID, value=resource_id['avatar'])
            e2.click()

            wrong_step = "点击头像控件"
            print_log("进入个人主页", end="-->")
            time.sleep(wait_time)
            
            
            
            e3 = driver.find_element(
                by=AppiumBy.ID, value=info_id['des']) # 点击更多
            e3.click()
            time.sleep(wait_time)
            get_xml(driver)
             
            info = analysis_info()
            if info == None:
                return False
            info_list.append(info)  # 信息添加

            wrong_step = "从个人主页返回文章"
            driver.find_element(
                by=AppiumBy.ID, value=resource_id['back1']).click()
            time.sleep(wait_time)

            wrong_step = "从文章返回首页"
            driver.find_element(
                by=AppiumBy.ID, value=resource_id['back2']).click()
            time.sleep(wait_time)
            print_log("退回主页")

        except:
            print_log(f"!e!某个控件不存在或页面出现其他异常,错误阶段{wrong_step}")
            get_xml(driver, "xml/wrong.xml")
            return False

    if fresh_page(driver):
        return True
    return False


def analysis_info(xml="xml/info.xml"):
    """解析xml文件中的信息"""
    #因为数据量不大,采用逐行匹配,而非xml树形解析

    info: dict[str, str] = {
    }
    with open(xml, "r", encoding="utf-8") as f:
        lines = f.readlines()

    for line in lines:
        for info_name, ids in info_id.items():
            if ids in line:
                padding = 6 if ids != info_id['xhs_id'] else 11  # 小红书id字段特殊处理
                start_num = line.find("text=\"") + padding
                end_num = line.find("\"", start_num)
                para = line[start_num:end_num]
                info[info_name] = para
                break
    if len(info) < 2:
        print_log("!e!未获取到博主信息,建议检查config文件")
        return None
    else:
        print_log("个人信息解析完成", end="-->")
        return info


def load_config(file="config.json"):
    """加载配置文件"""
    if not os.path.exists(file):
        print_log("config文件不存在")
        return None

    with open(file, 'r') as f:
        config = json.load(f)

    # print_log(config['device_config'].keys())
    if 'appium:noReset' in config['device_config'].keys():
        # boolean类型转换
        config['device_config']['appium:noReset'] = True if config['device_config']['appium:noReset'] == 'True' else False
    return config


def save_excel(info_list, file="output/info.xlsx"):
    columns_map = {
        'name_id': '用户名',
        'xhs_id': '小红书id',
        'des': '个人简介',
        'fans_num': '粉丝数',
        'all_likes_num': '点赞数',
        'article_likes_num': '前两篇点赞数',
    }
    if not os.path.exists("output/"):
        os.mkdir("output")
        print("output文件夹不存在,已自动创建")

    # 每日创建新的Sheet
    try:
        wb = openpyxl.load_workbook(file)
        print_log("已加载info.xlsx", end="-->")
    except:
        print_log("excel文件读取失败")

    date = time.strftime("%m-%d", time.localtime())
    if date not in wb.sheetnames:
        ws = wb.create_sheet(date, 0)
        print_log("新创建工作表", "-->")
        ws.append([v for _, v in columns_map.items()])

    # 激活当前的sheet
    wb.active = wb[date]
    ws = wb.active
    for info in info_list:
        meet = True  # 判断当前博主是否满足条件
        person_list = []  # 存放用户信息

        for k, _ in columns_map.items():
            if k in info.keys():
                # 添加异常判断
                person_list.append(info[k])
            else:
                meet = False
                break

        if not meet:
            continue

        try:
            fans_num = int(person_list[3]) # 粉丝数
            xhs_id = person_list[1] # 小红书id    
            if fans_num < 100:
                # 添加粉丝数等限制
                print_log("!w! 粉丝数不满足要求",end="")
                continue
            
            elif xhs_id in known_xhs_id:
                print_log("!w! 该用户已经存在",end="") 
            else:
                known_xhs_id.append(xhs_id)
                
        except:
            print_log("!w! 转换异常",end="")
            continue
        
        ws.append(person_list)

    wb.save("output/info.xlsx")
    print_log("新数据已保存至output/info.xlsx")


def init_driver(server_config, cap):
    driver1 = None
    server = f"{server_config['protocol']}://{server_config['host']}:{server_config['port']}{server_config['path']}"
    try:
        driver1 = webdriver.Remote(server, cap)
        driver1.implicitly_wait(3)  # 设置隐式等待时间为3秒.定位控件时间超过3秒 则报错
        print_log("appium服务连接成功")
    except:
        print_log("!e!appium服务启动失败,请检查appium服务是否启动")

    time.sleep(3)  # 等待手机响应
    return driver1


if __name__ == "__main__":
    now_time = time.strftime("%m-%d-%H-%M-%S", time.localtime())
    print_log(f"程序运行开始,时间:")

    config = load_config()
    if config is None:
        # 配置文件不存在
        print_log("!e! 配置文件不存在")
        sys.exit()
    cap = config['device_config']
    server_config = config['appium_server_config']
    
    
    
    global info_list
    info_list = []
    global resource_id
    global info_id
    resource_id = config['resource_id']
    info_id = config['info_id']
    
    with open('output/known_id.txt',"r",encoding='utf-8') as f:
        known_xhs_id= [line.strip("\n") for line in f.readlines()]    
    
    global driver
    driver = init_driver(server_config, cap)

    # 跳过广告
    # pass_ad(driver)
    # fresh_page(driver)
    
    error_num = 0
    max_erreor_num = 2
    plan_num = 400  # 计划获取的轮数
    times = 0  # 当前已经查询的数据量
    
    while times < plan_num:
        if not auto(driver):
            print_log(f"!e!程序出现异常,连续异常次数{error_num},重新开始执行")
            error_num += 1
            if error_num >= max_erreor_num:
                print_log("!e!程序连续多次出现异常,退出程序")
                break
            driver = init_driver(server_config, cap)
        
        else:
            error_num = 0
            times += 1  # 4是因为每次查询四个
            if len(info_list) > 0:  # 每四次写入excel一次
                save_excel(info_list)
                info_list = []
            else:
                print_log("!e!无新数据可以保存")
        
        try:        
            with open("output/known_id.txt","w",encoding="utf-8") as f:
                for id in known_xhs_id:
                    f.write(str(id)+"\n")
            print_log("known_id.txt已更新")
        except:
            print_log("!e! known_id.txt写入失败")
    
    print_log(f"程序运行结束,时间:")