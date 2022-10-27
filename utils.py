"""
此py文件用于处理
"""

import time
import os
import json
from tkinter import E
import openpyxl
import re
import traceback
    
def print_log(x,end="\n"):  
    """将日志保存到log文件夹下

    Args:
        x (str): 要打印的字符串
        now_time (str): 以日期为文件名
        end (str, optional): 结尾符. Defaults to "\n".
    """
    with open("tmp/.nowtime","r") as f:
        global now_time
        now_time = f.read()
    
    end = "  "+time.strftime("%H:%M:%S", time.localtime()
                             ) + end if end == "\n" else end
    date = time.strftime('%m-%d', time.localtime())+"/"
    if not os.path.exists("log/"+date):
        os.makedirs("log/"+date)

    with open("log/"+date+now_time+".log", "a+", encoding="utf-8") as f:
        f.write(str(x)+end)

    print(str(x), end=end)


def load_config(file="config.json"):
    """加载配置文件"""
    
    if not os.path.exists(file):
        print_log("config文件不存在")
        return None
    try:
        with open(file, 'r') as f:
            config = json.load(f)
    except:
        print_log("!e! config.json转换错误")
        return None

    if 'appium:noReset' in config['cap'].keys():
        # boolean类型转换
        config['cap']['appium:noReset'] = True if config['cap']['appium:noReset'] == 'True' or 'true' else False
    return config

def analysis_basic_info(info_id,known_xhs_id,xml="xml/info.xml"):
    """从xml文件中获取博主基本信息"""
    #因为数据量不大,采用逐行匹配,而非xml树形解析

    info: dict[str, str] = {
    }
    
    with open(xml, "r", encoding="utf-8") as f:
        lines = f.readlines()

    for line in lines:
        for info_name, ids in info_id.items():
            # 
            if ids in line:
                # 匹配是否是属于这一个字段的
                padding = 6 if ids != info_id['xhs_id'] else 11  # 小红书id字段特殊处理
                start_num = line.find("text=\"") + padding
                end_num = line.find("\"", start_num)
                para = line[start_num:end_num]
                info[info_name] = para
                
                # 进行筛选
                if ids == info_id['fans_num']:
                    print_log(f"粉丝{para}",end="-->")
                    # 粉丝数筛选
                    if "万" not in para:
                        try:
                            fans_num = int(para) # 粉丝数
                            if fans_num < 100:
                                # 添加粉丝数等限制
                                print_log("!w! 粉丝数不满足要求",end="-->")
                                return None
                        except:
                            print_log("!w! 粉丝数转换异常",end="-->")
                            return None
                
                elif ids == info_id['xhs_id']:
                    # 判断博主是否重复出现
                    if para in known_xhs_id:
                        print_log("博主已经存在",end="-->")
                        return None

                elif ids == info_id['des']:
                    # 从中提取邮箱号
                    info['email'] = get_mail(para)
                

    if len(info) < 2:
        print_log("!w!未获取到博主信息,可能出现直播或config文件异常")
        return None
    else:
        print_log("个人信息解析完成", end="-->")
        return info

def analysis_likes(likes,id,xml="xml/info.xml"):
    """从xml中获取点赞数量"""
    
    with open(xml, "r", encoding="utf-8") as f:
        lines = f.readlines()

    for line in lines:
        if id in line:
            start_num = line.find("text=\"") + 6
            end_num = line.find("\"", start_num)
            para = line[start_num:end_num]
            try:
                likes.append(int(para))
                print_log(f"点赞{para}",end="-->")
            except:
                if para == "赞":
                    likes.append(0)
                    print_log(f"点赞0",end="-->")
                elif "万" in para:
                    likes.append(int(float(para[:-1])*10000))
                else:
                    print_log("!e! 点赞数转换异常")
    return likes
            
def save_info(info, file="output/info.xlsx"):
    """将用户写入到excel中,并更新已知id"""
    
    columns_map = {
        'name_id': '用户名',
        'xhs_id': '小红书id',
        'des': '个人简介',
        'email': '初筛邮箱',
        'fans_num': '粉丝数',
        'all_likes_num': '总点赞数',
        'ave': '篇均点赞量',
    }
    if not os.path.exists("output/"):
        os.mkdir("output")
        print("output文件夹不存在,已自动创建")

    
    # 每日创建新的Sheet
    try:
        wb = openpyxl.load_workbook(file)
    except:
        wb = openpyxl.Workbook()
        print_log("新建output/info.xlsx", end="-->")

    date = time.strftime("%m-%d", time.localtime())
    
    if date not in wb.sheetnames:
        # 以当前日期新建sheet
        ws = wb.create_sheet(date, 0)
        print_log("新创建工作表", "-->")
        ws.append([v for _, v in columns_map.items()])

    # 激活当前的sheet
    wb.active = wb[date]
    ws = wb.active
    person_info_list = []  # 存放当前博主信息
    for k, _ in columns_map.items():
        if k in info.keys():
            # 添加异常判断
            person_info_list.append(info[k])
        else:
            print_log("!e!信息添加异常")
            return False
            
    ws.append(person_info_list)
    wb.save("output/info.xlsx")
    add_known_id(person_info_list[1])
    print_log("新信息已保存",end = "-->")
    return True

def load_known_id(file = 'output/known_id.txt'):
    if not os.path.exists("output/"):
        os.mkdir("output")
        print("output文件夹不存在,已自动创建")
        
    if not os.path.exists("output/known_id.txt"):
        with open(file, "w", encoding='utf-8'):
            pass

    with open(file, "r", encoding='utf-8') as f:
            known_xhs_id = [line.strip("\n") for line in f.readlines()]
    return known_xhs_id

def update_known_id(known_xhs_id,file = 'output/known_id.txt'):
    """更新已知博主的id"""
    try:
        with open(file, "w", encoding="utf-8") as f:
            for id in known_xhs_id:
                f.write(str(id)+"\n")
    except:
        print_log("!e! known_id.txt写入失败")
        
        
def add_known_id(new_id,file = 'output/known_id.txt'):
    """添加一行已知博主的id"""
    try:
        with open(file, "a+", encoding="utf-8") as f:
            f.write(str(new_id)+"\n")
        return True
    except:
        print_log("!e! known_id.txt写入失败")
        return False
        
def load_count(file = 'output/count.txt'):
    """加载统计数据"""
    if not os.path.exists("output/"):
        os.mkdir("output")
        print("output文件夹不存在,已自动创建")
        
    if not os.path.exists("output/count.txt"):
        with open(file, "w", encoding='utf-8') as f:
            # 初始化
            f.write("0\n0")

    with open(file, "r", encoding='utf-8') as f:
        lines = f.readlines()
    try:
        all_count = int(lines[0].strip("\n"))
        valid_count = int(lines[1])
        print_log(f"总收集量{all_count},有效收集量{valid_count}")
        return all_count,valid_count
    except:
        print_log("!e! 统计数据转换错误")
        return None,None
    
def update_count(all_count,valid_count,file = 'output/count.txt'):
    """更新统计数据"""
    try:
        with open(file, "w", encoding='utf-8') as f:
            f.write(str(all_count)+"\n")
            f.write(str(valid_count))
    except:
        print_log("!e! count.txt写入失败")


def get_mail(strs):
    """提取字符串中的邮箱"""
    p = re.compile(r'.*?(\w+([-+.]\w+)*@\w+([-.]\w+)*\.\w+([-.]\w+)*)',re.ASCII)
    for line in strs.splitlines():
        match = re.match(p,line)
        try :
            return match.group(1)
        except:
            pass
    return ''
    

if __name__ == "__main__":
    update_count(5,6)