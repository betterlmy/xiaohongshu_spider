import time
from appium import webdriver
from appium.webdriver.common.appiumby import AppiumBy
# from selenium.webdriver.support.wait import WebDriverWait
import os
import sys
import json
import openpyxl
# 控件id或XPATH路径存放字典
resource_id = {
    'avatar': "com.xingin.xhs:id/avatarLayout",
    'article': "com.xingin.xhs:id/cji",
    'video': "com.xingin.xhs:id/ckq",
    # 'fresh': "com.xingin.xhs:id/fo8",
    'fresh': "com.xingin.xhs:id/c8q",
    'back1': "com.xingin.xhs:id/egh",
    'back2': "com.xingin.xhs:id/oy",
    'pass': "com.xingin.xhs:id/f4e",
}

# 个人信息id
info_id = {
    'name_id': "com.xingin.xhs:id/egz",
    'xhs_id': "com.xingin.xhs:id/eh0",
    'des': "com.xingin.xhs:id/gjv",
    'fans_num': "com.xingin.xhs:id/b7d",
    'all_likes_num': "com.xingin.xhs:id/crd",
    'article_likes_num': "com.xingin.xhs:id/gf5"
}

info_list = []

now_time = ""


def print_log(x, end="\n"):
    """将日志保存到log文件夹下"""
    end = "  "+time.strftime("%H:%M:%S", time.localtime()) + end if end =="\n" else end
    date = time.strftime('%m-%d', time.localtime())+"/"
    if not os.path.exists("log/"+date):
        os.makedirs("log/"+date)

    with open("log/"+date+now_time+".log", "a+", encoding="utf-8") as f:
        f.write(x+end)

    print(x, end=end)


def pass_ad(driver):
    """跳过广告"""

    try:
        driver.find_element(by=AppiumBy.ID, value=resource_id["pass"]).click()
        time.sleep(1)
        print_log("跳过广告成功", end="-->")
    except:
        print_log("不存在广告", end="-->")


def get_xml(driver, xml="xml/info.xml"):
    """将页面资源写入到xml文件中"""

    if not os.path.exists("xml/"):
        os.mkdir("xml")
        print_log("xml文件夹不存在,已自动创建")

    with open(xml, "w", encoding="utf-8") as f:
        f.writelines(driver.page_source)
        print_log("xml文件写入成功", end="-->")


def contain_video(driver):
    """判断是否存在视频,存在返回True,不存在返回False"""

    try:
        driver.find_element(by=AppiumBy.ID, value=resource_id["video"])
        time.sleep(1)
        return True
    except:
        return False


def fresh_page(driver):
    """刷新页面"""

    try:
        # es = driver.find_elements(by=AppiumBy.CLASS_NAME, value='android.widget.TextView')
        # for e in es:
        #     if e.text == "发现":
        #         e.click()

        driver.find_element(by=AppiumBy.ID, value=resource_id['fresh']).click()
        # time.sleep(1)
        print_log("页面刷新成功", end="-->")
        return True
    except:
        print_log("不存在刷新按钮,刷新失败")
        return False


def auto(driver, max_time=8):
    """自动化主要流程"""

    fresh_time = 0  # 刷新次数
    while contain_video(driver):
        #如果包含了视频,直接fresh
        print_log("发现视频", end="-->")
        if fresh_page(driver):
            print_log("页面刷新一次", end="-->")
            fresh_time += 1
            if fresh_time > max_time:
                print_log("刷新次数过多仍未获取到信息,退出程序")
                return False
        else:
            return False

    print_log("未发现视频,开始获取信息")
    for index in range(4):
        # 每个阶段获取四个博主信息
        try:
            articles = driver.find_elements(
                by=AppiumBy.ID, value=resource_id["article"])
            articles[index].click()
            print_log(f"进入第{len(info_list)+1}篇文章", end="-->")
            time.sleep(1)
            e2 = driver.find_element(
                by=AppiumBy.ID, value=resource_id['avatar'])
            e2.click()
            print_log("进入个人主页", end="-->")
            time.sleep(1)

            get_xml(driver)
            info_list.append(analysis_info())
            driver.find_element(
                by=AppiumBy.ID, value=resource_id['back1']).click()
            time.sleep(1)
            driver.find_element(
                by=AppiumBy.ID, value=resource_id['back2']).click()
            time.sleep(1)
            print_log("退回主页")
        except:
            print_log("!e!某个控件不存在或页面出现其他异常")
            return False
    if fresh_page(driver):
        return True
    return False


def analysis_info(xml="xml/info.xml"):
    """解析xml文件中的信息"""
    #因为数据量不大,采用逐行匹配,而非xml树形解析

    info: dict[str, str] = {
    }
    with open(xml, "r", encoding="utf-8") as f:
        lines = f.readlines()

    for line in lines:
        for info_name, ids in info_id.items():
            if ids in line:
                padding = 6 if ids != "com.xingin.xhs:id/eh0" else 11  # 小红书id字段特殊处理
                start_num = line.find("text=\"") + padding
                end_num = line.find("\"", start_num)
                para = line[start_num:end_num]
                info[info_name] = para
                break

    print_log("个人信息解析完成", end="-->")
    return info


def load_config(file="config.json"):
    """加载配置文件"""
    if not os.path.exists(file):
        print_log("config文件不存在")
        return None

    with open(file, 'r') as f:
        config = json.load(f)

    # print_log(config['device_config'].keys())
    if 'appium:noReset' in config['device_config'].keys():
        # boolean类型转换
        config['device_config']['appium:noReset'] = True if config['device_config']['appium:noReset'] == 'True' else False
    return config


def save_excel(info_list, name="info.xlsx"):
    columns_map = {

        'name_id': '用户名',
        'xhs_id': '小红书id',
        'des': '个人简介',
        'fans_num': '粉丝数',
        'all_likes_num': '点赞数',
        'article_likes_num': '前两篇点赞数',
    }
    if not os.path.exists("output/"):
        os.mkdir("output")
        print("output文件夹不存在,已自动创建")

    # 每日创建新的Sheet
    try:
        wb = openpyxl.load_workbook("output/info.xlsx")
        print_log("已加载info.xlsx", end="-->")
    except:
        print_log("excel文件读取失败")

    date = time.strftime("%m-%d", time.localtime())
    if date not in wb.sheetnames:
        ws = wb.create_sheet(date, 0)
        print_log("新创建工作表", "-->")
        ws.append([v for _, v in columns_map.items()])

    wb.active = wb[date]
    ws = wb.active
    for info in info_list:
        meet = True # 判断当前博主是否满足条件
        person_list = [] # 存放用户信息
        
        for k, _ in columns_map.items():
            if k in info.keys():
                # 添加异常判断
                person_list.append(info[k])
            else:
                meet = False
                break
            
        if not meet:
            continue
        
        try:
            fans_num = int(person_list[3])
            if fans_num<100:
            # 添加粉丝数等限制
                continue
        except:
            print_log("!e! 转换异常")
            continue
            
        ws.append(person_list) 

    wb.save("output/info.xlsx")
    print_log("新数据已保存至output/info.xlsx")


driver = None

def init_driver():
    config = load_config()
    if config is None:
        # 配置文件不存在
        sys.exit()
    cap = config['device_config']
    server_config = config['appium_server_config']
    server = f"{server_config['protocol']}://{server_config['host']}:{server_config['port']}{server_config['path']}"
    try:
        global driver
        driver = webdriver.Remote(server, cap)
        driver.implicitly_wait(3)  # 设置隐式等待时间为3秒.定位控件时间超过3秒 则报错
        print_log("appium服务连接成功")
    except:
        print_log("!e!appium服务启动失败,请检查appium服务是否启动")
    

    time.sleep(3)  # 等待手机响应
    return driver

if __name__ == "__main__":
    now_time = time.strftime("%m-%d-%H-%M-%S", time.localtime())
    print_log(f"程序开始运行,开始时间:")

    driver = init_driver()
    
    # 跳过广告
    # pass_ad(driver)
    # fresh_page(driver)
    error_time = 0
    plan_num = 4 * 20  # 计划获取的博主数据量
    times = 0  # 当前已经查询的数据量
    while times < plan_num:
        if not auto(driver):
            print_log("!e!程序出现异常,重新开始执行")
            driver = init_driver()
            error_time += 1
            if error_time >= 4:
                print_log("!e!程序连续多次出现异常,退出程序")
                break
        error_time = 0
        times += 4  # 4是因为每次查询四个
        if len(info_list) > 0: # 每四次写入excel一次
            save_excel(info_list)
            info_list = []
        else:
            print_log("!e!无新数据可以保存")

